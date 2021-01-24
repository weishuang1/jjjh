#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2021/1/16 13:10 
# @Author : Lemon_Tricy
# @QQ: 2378807189
# Copyright：湖南省零檬信息技术有限公司
"""
代码做接口自动化测试，基本流程：
1、接口文档--接口测试用例--excel里 --ready
2、自动化excel表格读取测试用例数据--得到数据 -- ready
3、执行测试-- requests去执行--执行结果  --ready
4、执行结果 vs 预期结果 对比   ---是否通过？--最终结果
5、最终结果写入excel表格 ---ready--pass  failed

封装函数步骤：
1、实现功能的代码写出来 -- done
2、参数化？ -- 可变的数据--形参
3、返回值  -- 函数有没有需要给别人使用的东西

读取数据：第三方库  openpyxl--excel表格数据操作
1）读取  2）写入
excel 表格三大对象
1、文档--- 加载进来 -- 工作簿对象--load_workbook  ==最好放到同级 路径
2、表单对象：
3、单元格对象
注意： 写入生效的话--一定要保存文件,保存之前一定要关闭文件！
# 加载工作簿
wb = load_workbook("test_case_api.xlsx")
# 获取表单
sheet = wb["register"]
cell = sheet.cell(row=1,column=1)
# 获取单元格的值  对他重新赋值 --写入
cell.value = "用例编号"
wb.save("test_case_api.xlsx")

"""
import requests,pprint
import jsonpath
from openpyxl import load_workbook   # 导入指定的部分
# 读取数据的函数
def read_data(filename,sheetname):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    cases = []   # 空列表
    # 自动获取最大行号 列号
    max_row = sheet.max_row
    for i in range(2,max_row+1):  # 取头不取尾，—+1
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value,
        expected_result = sheet.cell(row=i,column=7).value)
        cases.append(case)
    return cases

# 写入结果函数
def write_result(filename,sheetname,final_result,row,column):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)

# 发送接口请求的函数
def api_func(url_api,data_api):
    head = {"X-Lemonban-Media-Type":"lemonban.v2"}
    response = requests.post(url=url_api,json=data_api,headers=head)
    return response.json()


